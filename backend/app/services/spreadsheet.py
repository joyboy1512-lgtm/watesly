import csv
import io
import json
import re
from pathlib import Path

XLSX_SIGNATURE = b"PK\x03\x04"
MAX_IMPORT_BYTES = 10 * 1024 * 1024

FIELD_ALIASES: dict[str, list[str]] = {
    "phone": [
        "phone",
        "external_address",
        "mobile",
        "tel",
        "telephone",
        "phone_number",
        "whatsapp",
        "رقم",
        "الهاتف",
        "جوال",
        "الجوال",
        "رقم_الهاتف",
        "رقم_الجوال",
    ],
    "name": ["name", "display_name", "customer", "client", "اسم", "الاسم", "اسم_العميل", "العميل"],
    "email": ["email", "e_mail", "mail", "بريد", "البريد", "email_address"],
    "language": ["language", "lang", "لغة", "اللغة"],
    "country_code": ["country_code", "country", "دولة", "رمز_الدولة"],
    "product_name": ["name", "product", "product_name", "service", "اسم", "الاسم", "منتج", "خدمة", "اسم_المنتج"],
    "product_type": ["product_type", "type", "kind", "نوع", "النوع"],
    "description": ["description", "desc", "details", "وصف", "الوصف", "تفاصيل"],
    "price": ["price", "amount", "cost", "سعر", "السعر", "التكلفة"],
    "currency": ["currency", "curr", "عملة", "العملة"],
    "price_type": ["price_type", "pricing", "نوع_السعر"],
    "sku": ["sku", "code", "product_code", "رمز", "كود"],
    "keywords": ["keywords", "tags", "search", "كلمات", "كلمات_البحث", "وسوم"],
    "specs": ["specs", "specifications", "مواصفات", "المواصفات"],
}


def normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("\ufeff", "")
    text = re.sub(r"\s+", "_", text)
    return text


def get_row_value(row: dict[str, str], field: str, aliases: list[str] | None = None) -> str:
    keys = aliases or FIELD_ALIASES.get(field, [field])
    for key in keys:
        normalized = normalize_header(key)
        if normalized in row and row[normalized].strip():
            return row[normalized].strip()
    return ""


def parse_spreadsheet(content: bytes, filename: str) -> list[dict[str, str]]:
    if len(content) > MAX_IMPORT_BYTES:
        raise ValueError("FILE_TOO_LARGE")

    suffix = Path(filename or "").suffix.lower()
    if suffix in {".xlsx", ".xlsm"} or content.startswith(XLSX_SIGNATURE):
        return _parse_xlsx(content)
    if suffix == ".csv" or suffix == ".txt" or _looks_like_text_table(content):
        return _parse_csv(content)
    raise ValueError("UNSUPPORTED_FILE_FORMAT")


def _looks_like_text_table(content: bytes) -> bool:
    if content.startswith(XLSX_SIGNATURE):
        return False
    sample = content[:4096]
    if b"\x00" in sample:
        return False
    try:
        sample.decode("utf-8")
    except UnicodeDecodeError:
        return False
    return b"," in sample or b";" in sample or b"\t" in sample


def _parse_csv(content: bytes) -> list[dict[str, str]]:
    text = content.decode("utf-8-sig")
    delimiter = ";" if text.count(";") > text.count(",") else ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    return [_normalize_row(row) for row in reader]


def _parse_xlsx(content: bytes) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return []

        headers = [normalize_header(cell) for cell in header_row]
        records: list[dict[str, str]] = []
        for row in rows_iter:
            if row is None or all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            record: dict[str, str] = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                value = row[index] if index < len(row) else None
                record[header] = "" if value is None else str(value).strip()
            records.append(record)
        return records
    finally:
        workbook.close()


def _normalize_row(row: dict[str, object]) -> dict[str, str]:
    return {normalize_header(key): "" if value is None else str(value).strip() for key, value in row.items()}


def parse_specs_value(raw: str) -> dict[str, str]:
    text = raw.strip()
    if not text:
        return {}
    if text.startswith("{"):
        try:
            parsed = json.loads(text)
            if isinstance(parsed, dict):
                return {str(k): str(v) for k, v in parsed.items()}
        except json.JSONDecodeError:
            pass

    specs: dict[str, str] = {}
    for chunk in re.split(r"[;\n|]", text):
        part = chunk.strip()
        if not part:
            continue
        if ":" in part:
            key, value = part.split(":", 1)
            specs[key.strip()] = value.strip()
        elif "=" in part:
            key, value = part.split("=", 1)
            specs[key.strip()] = value.strip()
    return specs


def collect_spec_columns(row: dict[str, str]) -> dict[str, str]:
    specs = parse_specs_value(get_row_value(row, "specs"))
    for key, value in row.items():
        if not value:
            continue
        if key.startswith("spec_"):
            specs[key.removeprefix("spec_")] = value
        elif key.startswith("مواصفة_"):
            specs[key.removeprefix("مواصفة_")] = value
    return specs


def normalize_product_type(raw: str) -> str:
    value = raw.strip().lower()
    if value in {"service", "خدمة", "services"}:
        return "service"
    return "product"


def normalize_price_type(raw: str) -> str:
    value = raw.strip().lower()
    if value in {"from", "يبدأ_من", "starts_from", "starting"}:
        return "from"
    if value in {"quote", "عرض_سعر", "quotation", "ask"}:
        return "quote"
    return "fixed"
