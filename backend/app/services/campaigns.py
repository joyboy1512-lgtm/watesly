from datetime import UTC, datetime
from uuid import UUID, uuid4

from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.campaign import Campaign, CampaignStatus
from app.models.campaign_recipient import CampaignRecipient, CampaignRecipientStatus
from app.models.contact import Contact
from app.models.conversation import Conversation, ConversationStatus
from app.models.message import Message, MessageDirection, MessageStatus, MessageType
from app.models.whatsapp_account import WhatsAppAccount
from app.models.whatsapp_template import TemplateStatus, WhatsAppTemplate
from app.schemas.campaign import CampaignCreateRequest
from app.services.outbox import add_outbox_event
from app.services.template_display import render_template_body_text

_RECIPIENT_TO_MESSAGE_STATUS = {
    CampaignRecipientStatus.SENT: MessageStatus.SENT,
    CampaignRecipientStatus.DELIVERED: MessageStatus.DELIVERED,
    CampaignRecipientStatus.READ: MessageStatus.READ,
    CampaignRecipientStatus.FAILED: MessageStatus.FAILED,
}

_ARCHIVABLE_STATUSES = {
    CampaignStatus.COMPLETED,
    CampaignStatus.COMPLETED_WITH_ERRORS,
    CampaignStatus.FAILED,
    CampaignStatus.CANCELLED,
}


async def create_campaign(db: AsyncSession, *, account_id: UUID, user_id: UUID, payload: CampaignCreateRequest) -> Campaign:
    wa = await db.get(WhatsAppAccount, payload.whatsapp_account_id)
    template = await db.get(WhatsAppTemplate, payload.template_id)
    if wa is None or wa.account_id != account_id:
        raise ValueError("INVALID_WHATSAPP_ACCOUNT")
    if wa.organization_id != payload.organization_id:
        raise ValueError("ORGANIZATION_MISMATCH")
    if template is None or template.account_id != account_id:
        raise ValueError("INVALID_TEMPLATE")
    if template.whatsapp_account_id != wa.id:
        raise ValueError("TEMPLATE_ACCOUNT_MISMATCH")
    if template.status != TemplateStatus.APPROVED:
        raise ValueError("TEMPLATE_NOT_APPROVED")

    contact_ids = [r.contact_id for r in payload.recipients]
    if len(set(contact_ids)) != len(contact_ids):
        raise ValueError("DUPLICATE_RECIPIENT")
    contact_query = select(Contact.id, Contact.marketing_opt_in).where(
        Contact.account_id == account_id,
        Contact.organization_id == payload.organization_id,
        Contact.deleted_at.is_(None),
        Contact.id.in_(contact_ids),
    )
    result = await db.execute(contact_query)
    rows = {row[0]: row[1] for row in result.all()}
    if set(rows.keys()) != set(contact_ids):
        raise ValueError("INVALID_RECIPIENT")

    recipients = payload.recipients
    if payload.exclude_marketing_opt_out:
        recipients = [item for item in recipients if rows.get(item.contact_id) is not False]
        if not recipients:
            raise ValueError("ALL_RECIPIENTS_OPTED_OUT")

    campaign = Campaign(
        account_id=account_id,
        organization_id=payload.organization_id,
        whatsapp_account_id=payload.whatsapp_account_id,
        template_id=payload.template_id,
        created_by_user_id=user_id,
        name=payload.name,
        status=CampaignStatus.DRAFT,
        scheduled_at=payload.scheduled_at,
        max_recipients=min(max(len(recipients), 1), 10000),
        requires_approval=True,
        include_opt_out_option=payload.include_opt_out_option,
    )
    db.add(campaign)
    await db.flush()
    db.add_all([CampaignRecipient(
        campaign_id=campaign.id,
        contact_id=item.contact_id,
        template_parameters=item.template_parameters,
    ) for item in recipients])
    await add_outbox_event(db, account_id=account_id, event_type="campaign.created", aggregate_type="campaign", aggregate_id=str(campaign.id), payload={"campaign_id": str(campaign.id)})
    await db.commit()
    await db.refresh(campaign)
    return campaign


async def list_campaigns(
    db: AsyncSession,
    account_id: UUID,
    *,
    limit: int = 100,
    include_archived: bool = False,
    archived_only: bool = False,
) -> list[Campaign]:
    query = (
        select(Campaign)
        .where(Campaign.account_id == account_id)
        .order_by(Campaign.created_at.desc())
        .limit(min(max(limit, 1), 200))
    )
    if archived_only:
        query = query.where(Campaign.archived_at.is_not(None))
    elif not include_archived:
        query = query.where(Campaign.archived_at.is_(None))
    result = await db.execute(query)
    return list(result.scalars().all())


async def list_campaigns_with_reports(
    db: AsyncSession,
    account_id: UUID,
    *,
    limit: int = 100,
    include_archived: bool = False,
    archived_only: bool = False,
) -> list[dict]:
    campaigns = await list_campaigns(
        db,
        account_id,
        limit=limit,
        include_archived=include_archived,
        archived_only=archived_only,
    )
    items: list[dict] = []
    for campaign in campaigns:
        report = await get_campaign_report(
            db, account_id=account_id, campaign_id=campaign.id
        )
        status = campaign.status.value if hasattr(campaign.status, "value") else str(campaign.status)
        items.append({
            "id": campaign.id,
            "organization_id": campaign.organization_id,
            "whatsapp_account_id": campaign.whatsapp_account_id,
            "template_id": campaign.template_id,
            "name": campaign.name,
            "status": status,
            "scheduled_at": campaign.scheduled_at,
            "started_at": campaign.started_at,
            "completed_at": campaign.completed_at,
            "include_opt_out_option": campaign.include_opt_out_option,
            "archived_at": campaign.archived_at,
            "report": report,
        })
    return items


async def get_campaign(db: AsyncSession, *, account_id: UUID, campaign_id: UUID) -> Campaign:
    campaign = await db.get(Campaign, campaign_id)
    if campaign is None or campaign.account_id != account_id:
        raise ValueError("CAMPAIGN_NOT_FOUND")
    return campaign


async def recipient_count(db: AsyncSession, campaign_id: UUID) -> int:
    return int((await db.scalar(select(func.count(CampaignRecipient.id)).where(CampaignRecipient.campaign_id == campaign_id))) or 0)


async def approve_campaign(db: AsyncSession, *, account_id: UUID, campaign_id: UUID, user_id: UUID) -> Campaign:
    campaign = await get_campaign(db, account_id=account_id, campaign_id=campaign_id)
    if campaign.status not in {CampaignStatus.DRAFT, CampaignStatus.PAUSED}:
        raise ValueError("CAMPAIGN_CANNOT_BE_APPROVED")
    count = await recipient_count(db, campaign.id)
    if count > campaign.max_recipients:
        raise ValueError("RECIPIENT_LIMIT_EXCEEDED")
    campaign.approved_by_user_id = user_id
    campaign.approved_at = datetime.now(UTC)
    await db.commit(); await db.refresh(campaign)
    return campaign


async def prepare_campaign_start(db: AsyncSession, *, account_id: UUID, campaign_id: UUID) -> Campaign:
    campaign = (await db.execute(select(Campaign).where(Campaign.id==campaign_id,Campaign.account_id==account_id).with_for_update())).scalar_one_or_none()
    if campaign is None: raise ValueError("CAMPAIGN_NOT_FOUND")
    if campaign.status not in {CampaignStatus.DRAFT, CampaignStatus.SCHEDULED, CampaignStatus.PAUSED}:
        raise ValueError("CAMPAIGN_CANNOT_START")
    if campaign.requires_approval and campaign.approved_at is None:
        raise ValueError("CAMPAIGN_APPROVAL_REQUIRED")
    if await recipient_count(db, campaign.id) > campaign.max_recipients:
        raise ValueError("RECIPIENT_LIMIT_EXCEEDED")
    campaign.status = CampaignStatus.SCHEDULED
    campaign.execution_token = uuid4()
    campaign.active_task_id = None
    campaign.paused_at = None
    campaign.cancelled_reason = None
    await db.commit(); await db.refresh(campaign)
    return campaign


async def pause_campaign(db: AsyncSession, *, account_id: UUID, campaign_id: UUID) -> Campaign:
    campaign = await get_campaign(db, account_id=account_id, campaign_id=campaign_id)
    if campaign.status not in {CampaignStatus.SCHEDULED, CampaignStatus.RUNNING}:
        raise ValueError("CAMPAIGN_CANNOT_PAUSE")
    campaign.status = CampaignStatus.PAUSED
    campaign.paused_at = datetime.now(UTC)
    await db.commit(); await db.refresh(campaign)
    return campaign


async def cancel_campaign(db: AsyncSession, *, account_id: UUID, campaign_id: UUID, reason: str) -> Campaign:
    campaign = await get_campaign(db, account_id=account_id, campaign_id=campaign_id)
    if campaign.status in {CampaignStatus.COMPLETED, CampaignStatus.CANCELLED}:
        raise ValueError("CAMPAIGN_CANNOT_CANCEL")
    campaign.status = CampaignStatus.CANCELLED
    campaign.cancelled_reason = reason[:2000]
    campaign.completed_at = datetime.now(UTC)
    await db.commit(); await db.refresh(campaign)
    return campaign


async def archive_campaign(db: AsyncSession, *, account_id: UUID, campaign_id: UUID) -> Campaign:
    campaign = await get_campaign(db, account_id=account_id, campaign_id=campaign_id)
    if campaign.status not in _ARCHIVABLE_STATUSES:
        raise ValueError("CAMPAIGN_CANNOT_ARCHIVE")
    if campaign.archived_at is None:
        campaign.archived_at = datetime.now(UTC)
        await add_outbox_event(
            db,
            account_id=account_id,
            event_type="campaign.archived",
            aggregate_type="campaign",
            aggregate_id=str(campaign.id),
            payload={"campaign_id": str(campaign.id)},
        )
        await db.commit()
        await db.refresh(campaign)
    return campaign


async def unarchive_campaign(db: AsyncSession, *, account_id: UUID, campaign_id: UUID) -> Campaign:
    campaign = await get_campaign(db, account_id=account_id, campaign_id=campaign_id)
    if campaign.archived_at is None:
        raise ValueError("CAMPAIGN_NOT_ARCHIVED")
    campaign.archived_at = None
    await add_outbox_event(
        db,
        account_id=account_id,
        event_type="campaign.unarchived",
        aggregate_type="campaign",
        aggregate_id=str(campaign.id),
        payload={"campaign_id": str(campaign.id)},
    )
    await db.commit()
    await db.refresh(campaign)
    return campaign


async def delete_draft_campaign(db: AsyncSession, *, account_id: UUID, campaign_id: UUID) -> None:
    campaign = await get_campaign(db, account_id=account_id, campaign_id=campaign_id)
    if campaign.status != CampaignStatus.DRAFT:
        raise ValueError("CAMPAIGN_CANNOT_DELETE")
    sent_count = int(
        (
            await db.scalar(
                select(func.count(CampaignRecipient.id)).where(
                    CampaignRecipient.campaign_id == campaign.id,
                    CampaignRecipient.external_message_id.is_not(None),
                )
            )
        )
        or 0
    )
    if sent_count > 0:
        raise ValueError("CAMPAIGN_CANNOT_DELETE")
    follow_up_count = int(
        (
            await db.scalar(
                select(func.count(Campaign.id)).where(Campaign.parent_campaign_id == campaign.id)
            )
        )
        or 0
    )
    if follow_up_count > 0:
        raise ValueError("CAMPAIGN_HAS_FOLLOW_UPS")
    await add_outbox_event(
        db,
        account_id=account_id,
        event_type="campaign.deleted",
        aggregate_type="campaign",
        aggregate_id=str(campaign.id),
        payload={"campaign_id": str(campaign.id), "name": campaign.name},
    )
    await db.delete(campaign)
    await db.commit()


async def get_campaign_report(db: AsyncSession, *, account_id: UUID, campaign_id: UUID) -> dict:
    await get_campaign(db, account_id=account_id, campaign_id=campaign_id)
    result = await db.execute(select(CampaignRecipient.status, func.count(CampaignRecipient.id)).where(CampaignRecipient.campaign_id == campaign_id).group_by(CampaignRecipient.status))
    counts: dict[CampaignRecipientStatus, int] = {}
    for status, count in result.all():
        key = status if isinstance(status, CampaignRecipientStatus) else CampaignRecipientStatus(str(status))
        counts[key] = counts.get(key, 0) + int(count)
    total = sum(counts.values()); delivered = counts.get(CampaignRecipientStatus.DELIVERED, 0); read = counts.get(CampaignRecipientStatus.READ, 0); sent = counts.get(CampaignRecipientStatus.SENT, 0)
    return {"total": total, "pending": counts.get(CampaignRecipientStatus.PENDING, 0), "queued": counts.get(CampaignRecipientStatus.QUEUED, 0), "sent": sent, "delivered": delivered, "read": read, "failed": counts.get(CampaignRecipientStatus.FAILED, 0), "skipped": counts.get(CampaignRecipientStatus.SKIPPED, 0), "delivery_rate": round(((delivered + read) / max(sent + delivered + read, 1)) * 100, 2), "read_rate": round((read / max(delivered + read, 1)) * 100, 2)}


async def list_campaign_recipients(
    db: AsyncSession, *, account_id: UUID, campaign_id: UUID
) -> list[dict]:
    await get_campaign(db, account_id=account_id, campaign_id=campaign_id)
    rows = await db.execute(
        select(CampaignRecipient, Contact)
        .join(Contact, Contact.id == CampaignRecipient.contact_id)
        .where(CampaignRecipient.campaign_id == campaign_id)
        .order_by(CampaignRecipient.status.asc(), Contact.external_address.asc())
    )
    items: list[dict] = []
    for recipient, contact in rows.all():
        status = recipient.status.value if hasattr(recipient.status, "value") else str(recipient.status)
        items.append({
            "contact_id": str(contact.id),
            "display_name": contact.display_name,
            "phone": contact.external_address,
            "status": status,
            "error_message": recipient.error_message,
        })
    return items


async def export_campaign_recipients_csv(
    db: AsyncSession, *, account_id: UUID, campaign_id: UUID
) -> str:
    import csv
    import io

    items = await list_campaign_recipients(db, account_id=account_id, campaign_id=campaign_id)
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["name", "phone", "status", "error"])
    for item in items:
        writer.writerow([
            item["display_name"] or "",
            item["phone"],
            item["status"],
            item["error_message"] or "",
        ])
    return buffer.getvalue()


async def export_campaign_recipients_xlsx(
    db: AsyncSession, *, account_id: UUID, campaign_id: UUID
) -> bytes:
    import io

    from openpyxl import Workbook

    items = await list_campaign_recipients(db, account_id=account_id, campaign_id=campaign_id)
    workbook = Workbook()
    workbook.remove(workbook.active)

    summary = await get_campaign_report(db, account_id=account_id, campaign_id=campaign_id)
    summary_sheet = workbook.create_sheet("ملخص")
    summary_sheet.append(["المؤشر", "القيمة"])
    for key, value in summary.items():
        summary_sheet.append([key, value])

    detail = workbook.create_sheet("كل الأرقام")
    detail.append(["الاسم", "الرقم", "الحالة", "سبب الفشل"])
    for item in items:
        detail.append([
            item["display_name"] or "",
            item["phone"],
            item["status"],
            item["error_message"] or "",
        ])

    for status in ("sent", "delivered", "read", "failed", "pending"):
        group = [i for i in items if i["status"] == status]
        if not group:
            continue
        sheet = workbook.create_sheet(status[:31])
        sheet.append(["الاسم", "الرقم", "سبب الفشل"])
        for item in group:
            sheet.append([item["display_name"] or "", item["phone"], item["error_message"] or ""])

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


async def record_campaign_outbound_message(
    db: AsyncSession,
    *,
    campaign: Campaign,
    recipient: CampaignRecipient,
    contact: Contact,
    wa: WhatsAppAccount,
    template: WhatsAppTemplate,
    to_address: str,
    external_message_id: str,
    send_components: list,
    recipient_status: CampaignRecipientStatus = CampaignRecipientStatus.SENT,
) -> Message | None:
    """Persist a successful campaign send in inbox messages/conversations."""
    if not external_message_id:
        return None

    existing = (
        await db.execute(select(Message).where(Message.external_message_id == external_message_id))
    ).scalar_one_or_none()
    if existing is not None:
        mapped = _RECIPIENT_TO_MESSAGE_STATUS.get(recipient_status)
        if mapped and existing.status != mapped:
            existing.status = mapped
            if recipient.error_message and recipient_status == CampaignRecipientStatus.FAILED:
                payload = existing.provider_payload if isinstance(existing.provider_payload, dict) else {}
                existing.provider_payload = {**payload, "delivery_error": recipient.error_message}
        return existing

    conv_result = await db.execute(
        select(Conversation).where(
            Conversation.contact_id == contact.id,
            Conversation.channel_id == wa.channel_id,
            Conversation.deleted_at.is_(None),
            Conversation.status.in_([ConversationStatus.OPEN, ConversationStatus.PENDING]),
        )
    )
    conversation = conv_result.scalars().first()
    if conversation is None:
        conversation = Conversation(
            account_id=campaign.account_id,
            organization_id=campaign.organization_id,
            channel_id=wa.channel_id,
            contact_id=contact.id,
            status=ConversationStatus.OPEN,
        )
        db.add(conversation)
        await db.flush()

    message_status = _RECIPIENT_TO_MESSAGE_STATUS.get(recipient_status, MessageStatus.SENT)
    message = Message(
        account_id=campaign.account_id,
        organization_id=campaign.organization_id,
        channel_id=wa.channel_id,
        contact_id=contact.id,
        conversation_id=conversation.id,
        external_message_id=external_message_id,
        direction=MessageDirection.OUTBOUND,
        type=MessageType.TEMPLATE,
        from_address=wa.display_phone_number,
        to_address=to_address,
        text_body=render_template_body_text(template.components, fallback=template.body_text),
        provider_payload={
            "template_name": template.name,
            "language_code": template.language,
            "campaign_id": str(campaign.id),
            "campaign_recipient_id": str(recipient.id),
            "send_components": send_components,
            "components": template.components,
        },
        status=message_status,
    )
    db.add(message)
    now = datetime.now(UTC)
    conversation.last_message_at = now
    await db.flush()
    return message


async def backfill_campaign_inbox_messages(db: AsyncSession, *, account_id: UUID, limit: int = 500) -> int:
    """Create inbox messages for campaign sends that were not previously recorded."""
    rows = (
        await db.execute(
            select(CampaignRecipient, Campaign, Contact, WhatsAppAccount, WhatsAppTemplate)
            .join(Campaign, Campaign.id == CampaignRecipient.campaign_id)
            .join(Contact, Contact.id == CampaignRecipient.contact_id)
            .join(WhatsAppAccount, WhatsAppAccount.id == Campaign.whatsapp_account_id)
            .join(WhatsAppTemplate, WhatsAppTemplate.id == Campaign.template_id)
            .outerjoin(Message, Message.external_message_id == CampaignRecipient.external_message_id)
            .where(
                Campaign.account_id == account_id,
                CampaignRecipient.external_message_id.is_not(None),
                Message.id.is_(None),
                CampaignRecipient.status.in_([
                    CampaignRecipientStatus.SENT,
                    CampaignRecipientStatus.DELIVERED,
                    CampaignRecipientStatus.READ,
                    CampaignRecipientStatus.FAILED,
                ]),
            )
            .order_by(CampaignRecipient.last_attempt_at.desc())
            .limit(min(max(limit, 1), 2000))
        )
    ).all()

    created = 0
    for recipient, campaign, contact, wa, template in rows:
        from app.services.phone_normalize import normalize_whatsapp_phone

        dial = {"KW": "965", "SA": "966", "AE": "971", "QA": "974", "BH": "973", "OM": "968"}.get(
            (contact.country_code or "KW").upper(), "965"
        )
        to_address = normalize_whatsapp_phone(contact.external_address, country_code=dial) or contact.external_address
        components = recipient.template_parameters or []
        message = await record_campaign_outbound_message(
            db,
            campaign=campaign,
            recipient=recipient,
            contact=contact,
            wa=wa,
            template=template,
            to_address=to_address,
            external_message_id=recipient.external_message_id or "",
            send_components=components if isinstance(components, list) else [],
            recipient_status=recipient.status,
        )
        if message is not None:
            created += 1
    if created:
        await db.commit()
    return created
