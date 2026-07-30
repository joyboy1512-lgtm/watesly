from datetime import UTC, datetime, timedelta
from uuid import UUID

from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.campaign import Campaign
from app.models.campaign_recipient import CampaignRecipient, CampaignRecipientStatus
from app.models.catalog_product import CatalogProduct
from app.models.contact import Contact
from app.models.conversation import Conversation, ConversationStatus
from app.models.message import Message, MessageDirection


def _contact_row(contact: Contact) -> dict:
    return {
        "id": str(contact.id),
        "display_name": contact.display_name,
        "phone": contact.external_address,
        "email": contact.email,
        "country_code": contact.country_code,
        "created_at": contact.created_at.isoformat() if contact.created_at else None,
    }


async def reports_overview(db: AsyncSession, *, account_id: UUID, days: int = 30) -> dict:
    since = datetime.now(UTC) - timedelta(days=days)

    total_contacts = int(
        (await db.scalar(select(func.count(Contact.id)).where(
            Contact.account_id == account_id, Contact.deleted_at.is_(None)
        )))
        or 0
    )
    new_contacts = int(
        (await db.scalar(select(func.count(Contact.id)).where(
            Contact.account_id == account_id,
            Contact.deleted_at.is_(None),
            Contact.created_at >= since,
        )))
        or 0
    )
    with_name = int(
        (await db.scalar(select(func.count(Contact.id)).where(
            Contact.account_id == account_id,
            Contact.deleted_at.is_(None),
            Contact.display_name.is_not(None),
            func.trim(Contact.display_name) != "",
        )))
        or 0
    )
    engagement = await engagement_report(db, account_id=account_id, days=days, limit=0)
    campaign_count = int(
        (await db.scalar(select(func.count(Campaign.id)).where(
            Campaign.account_id == account_id, Campaign.created_at >= since
        )))
        or 0
    )
    open_conversations = int(
        (await db.scalar(select(func.count(Conversation.id)).where(
            Conversation.account_id == account_id,
            Conversation.deleted_at.is_(None),
            Conversation.status == ConversationStatus.OPEN,
        )))
        or 0
    )
    conv_summary = await conversations_report(db, account_id=account_id, days=days, limit=1)
    inactive_summary = await inactivity_report(db, account_id=account_id, inactive_days=days, limit=50)
    return {
        "period_days": days,
        "total_contacts": total_contacts,
        "new_contacts": new_contacts,
        "contacts_with_name": with_name,
        "contacts_without_name": max(total_contacts - with_name, 0),
        "two_way_engaged": engagement["summary"]["two_way_engaged"],
        "waiting_team_reply": engagement["summary"]["waiting_team_reply"],
        "waiting_customer_reply": engagement["summary"]["waiting_customer_reply"],
        "no_interaction": engagement["summary"]["no_interaction"],
        "campaigns_in_period": campaign_count,
        "open_conversations": open_conversations,
        "sla_breaches": conv_summary["summary"]["sla_breaches"],
        "inactive_contacts": inactive_summary["summary"]["total_inactive"],
    }


async def customer_report(db: AsyncSession, *, account_id: UUID, days: int = 30, limit: int = 50) -> dict:
    since = datetime.now(UTC) - timedelta(days=days)

    total_contacts = int(
        (await db.scalar(select(func.count(Contact.id)).where(
            Contact.account_id == account_id, Contact.deleted_at.is_(None)
        )))
        or 0
    )
    new_contacts = int(
        (await db.scalar(select(func.count(Contact.id)).where(
            Contact.account_id == account_id,
            Contact.deleted_at.is_(None),
            Contact.created_at >= since,
        )))
        or 0
    )
    with_email = int(
        (await db.scalar(select(func.count(Contact.id)).where(
            Contact.account_id == account_id,
            Contact.deleted_at.is_(None),
            Contact.email.is_not(None),
            func.trim(Contact.email) != "",
        )))
        or 0
    )
    with_conversations = int(
        (await db.scalar(
            select(func.count(func.distinct(Contact.id)))
            .join(Conversation, Conversation.contact_id == Contact.id)
            .where(
                Contact.account_id == account_id,
                Contact.deleted_at.is_(None),
                Conversation.deleted_at.is_(None),
            )
        ))
        or 0
    )

    country_rows = (
        await db.execute(
            select(Contact.country_code, func.count(Contact.id).label("count"))
            .where(Contact.account_id == account_id, Contact.deleted_at.is_(None))
            .group_by(Contact.country_code)
            .order_by(func.count(Contact.id).desc())
            .limit(10)
        )
    ).all()

    recent = list(
        (
            await db.execute(
                select(Contact)
                .where(Contact.account_id == account_id, Contact.deleted_at.is_(None))
                .order_by(Contact.created_at.desc())
                .limit(limit)
            )
        ).scalars().all()
    )

    return {
        "period_days": days,
        "summary": {
            "total_contacts": total_contacts,
            "new_contacts": new_contacts,
            "with_email": with_email,
            "with_conversations": with_conversations,
            "without_conversations": max(total_contacts - with_conversations, 0),
        },
        "by_country": [
            {"country_code": row.country_code or "—", "count": int(row.count)} for row in country_rows
        ],
        "recent_contacts": [_contact_row(item) for item in recent],
    }


async def names_report(db: AsyncSession, *, account_id: UUID, limit: int = 50) -> dict:
    total_contacts = int(
        (await db.scalar(select(func.count(Contact.id)).where(
            Contact.account_id == account_id, Contact.deleted_at.is_(None)
        )))
        or 0
    )
    with_name = int(
        (await db.scalar(select(func.count(Contact.id)).where(
            Contact.account_id == account_id,
            Contact.deleted_at.is_(None),
            Contact.display_name.is_not(None),
            func.trim(Contact.display_name) != "",
        )))
        or 0
    )

    duplicate_rows = (
        await db.execute(
            select(Contact.display_name, func.count(Contact.id).label("count"))
            .where(
                Contact.account_id == account_id,
                Contact.deleted_at.is_(None),
                Contact.display_name.is_not(None),
                func.trim(Contact.display_name) != "",
            )
            .group_by(Contact.display_name)
            .having(func.count(Contact.id) > 1)
            .order_by(func.count(Contact.id).desc())
            .limit(limit)
        )
    ).all()

    missing_names = list(
        (
            await db.execute(
                select(Contact)
                .where(
                    Contact.account_id == account_id,
                    Contact.deleted_at.is_(None),
                    (Contact.display_name.is_(None)) | (func.trim(Contact.display_name) == ""),
                )
                .order_by(Contact.created_at.desc())
                .limit(limit)
            )
        ).scalars().all()
    )

    named_contacts = list(
        (
            await db.execute(
                select(Contact)
                .where(
                    Contact.account_id == account_id,
                    Contact.deleted_at.is_(None),
                    Contact.display_name.is_not(None),
                    func.trim(Contact.display_name) != "",
                )
                .order_by(Contact.display_name.asc())
                .limit(limit)
            )
        ).scalars().all()
    )

    return {
        "summary": {
            "total_contacts": total_contacts,
            "with_name": with_name,
            "without_name": max(total_contacts - with_name, 0),
            "duplicate_name_groups": len(duplicate_rows),
        },
        "duplicate_names": [
            {"name": row.display_name, "count": int(row.count)} for row in duplicate_rows
        ],
        "missing_names": [_contact_row(item) for item in missing_names],
        "named_contacts": [_contact_row(item) for item in named_contacts],
    }


async def _message_counts_by_contact(db: AsyncSession, *, account_id: UUID, since: datetime) -> dict[UUID, dict[str, int]]:
    rows = (
        await db.execute(
            select(
                Message.contact_id,
                Message.direction,
                func.count(Message.id).label("count"),
            )
            .where(
                Message.account_id == account_id,
                Message.contact_id.is_not(None),
                Message.created_at >= since,
            )
            .group_by(Message.contact_id, Message.direction)
        )
    ).all()

    stats: dict[UUID, dict[str, int]] = {}
    for row in rows:
        if row.contact_id is None:
            continue
        bucket = stats.setdefault(row.contact_id, {"inbound": 0, "outbound": 0})
        direction = row.direction.value if hasattr(row.direction, "value") else str(row.direction)
        bucket[direction] = int(row.count)
    return stats


async def engagement_report(
    db: AsyncSession,
    *,
    account_id: UUID,
    days: int = 30,
    limit: int = 50,
) -> dict:
    since = datetime.now(UTC) - timedelta(days=days)
    stats = await _message_counts_by_contact(db, account_id=account_id, since=since)

    two_way_ids: list[UUID] = []
    waiting_team_ids: list[UUID] = []
    waiting_customer_ids: list[UUID] = []
    no_interaction_ids: list[UUID] = []

    for contact_id, counts in stats.items():
        inbound = counts.get("inbound", 0)
        outbound = counts.get("outbound", 0)
        if inbound > 0 and outbound > 0:
            two_way_ids.append(contact_id)
        elif inbound > 0 and outbound == 0:
            waiting_team_ids.append(contact_id)
        elif outbound > 0 and inbound == 0:
            waiting_customer_ids.append(contact_id)

    all_contacts = list(
        (
            await db.execute(
                select(Contact.id).where(Contact.account_id == account_id, Contact.deleted_at.is_(None))
            )
        ).scalars().all()
    )
    messaged_ids = set(stats.keys())
    no_interaction_ids = [cid for cid in all_contacts if cid not in messaged_ids]

    async def load_contacts(ids: list[UUID]) -> list[dict]:
        if not ids or limit == 0:
            return []
        items = list(
            (
                await db.execute(
                    select(Contact)
                    .where(Contact.id.in_(ids[:limit]))
                    .order_by(Contact.updated_at.desc())
                )
            ).scalars().all()
        )
        return [_contact_row(item) for item in items]

    open_waiting_team = int(
        (await db.scalar(
            select(func.count(Conversation.id)).where(
                Conversation.account_id == account_id,
                Conversation.deleted_at.is_(None),
                Conversation.status.in_(["open", "pending"]),
                Conversation.first_response_at.is_(None),
            )
        ))
        or 0
    )

    return {
        "period_days": days,
        "summary": {
            "two_way_engaged": len(two_way_ids),
            "waiting_team_reply": len(waiting_team_ids),
            "waiting_customer_reply": len(waiting_customer_ids),
            "no_interaction": len(no_interaction_ids),
            "open_conversations_waiting_team": open_waiting_team,
        },
        "two_way_engaged": await load_contacts(two_way_ids),
        "waiting_team_reply": await load_contacts(waiting_team_ids),
        "waiting_customer_reply": await load_contacts(waiting_customer_ids),
        "no_interaction": await load_contacts(no_interaction_ids[:limit]),
    }


async def campaigns_report(db: AsyncSession, *, account_id: UUID, days: int = 30, limit: int = 50) -> dict:
    from app.services.campaigns import get_campaign_report

    since = datetime.now(UTC) - timedelta(days=days)
    campaigns = list(
        (
            await db.execute(
                select(Campaign)
                .where(Campaign.account_id == account_id, Campaign.created_at >= since)
                .order_by(Campaign.created_at.desc())
                .limit(limit)
            )
        ).scalars().all()
    )

    rows = []
    totals = {"campaigns": 0, "recipients": 0, "sent": 0, "delivered": 0, "read": 0, "failed": 0}
    for campaign in campaigns:
        stats = await get_campaign_report(db, account_id=account_id, campaign_id=campaign.id)
        status = campaign.status.value if hasattr(campaign.status, "value") else str(campaign.status)
        rows.append({
            "id": str(campaign.id),
            "name": campaign.name,
            "status": status,
            "created_at": campaign.created_at.isoformat() if campaign.created_at else None,
            **stats,
        })
        totals["campaigns"] += 1
        totals["recipients"] += stats["total"]
        totals["sent"] += stats["sent"]
        totals["delivered"] += stats["delivered"]
        totals["read"] += stats["read"]
        totals["failed"] += stats["failed"]

    delivery_base = max(totals["sent"] + totals["delivered"] + totals["read"], 1)
    read_base = max(totals["delivered"] + totals["read"], 1)
    totals["delivery_rate"] = round(((totals["delivered"] + totals["read"]) / delivery_base) * 100, 1)
    totals["read_rate"] = round((totals["read"] / read_base) * 100, 1)

    return {"period_days": days, "summary": totals, "campaigns": rows}


async def conversations_report(
    db: AsyncSession, *, account_id: UUID, days: int = 30, limit: int = 50, sla_minutes: int = 15
) -> dict:
    since = datetime.now(UTC) - timedelta(days=days)
    sla_delta = timedelta(minutes=sla_minutes)

    base = Conversation.account_id == account_id, Conversation.deleted_at.is_(None)
    open_count = int(
        (await db.scalar(select(func.count(Conversation.id)).where(*base, Conversation.status == ConversationStatus.OPEN)))
        or 0
    )
    pending_count = int(
        (await db.scalar(select(func.count(Conversation.id)).where(*base, Conversation.status == ConversationStatus.PENDING)))
        or 0
    )
    closed_count = int(
        (await db.scalar(select(func.count(Conversation.id)).where(*base, Conversation.status == ConversationStatus.CLOSED)))
        or 0
    )
    period_count = int(
        (await db.scalar(select(func.count(Conversation.id)).where(*base, Conversation.created_at >= since)))
        or 0
    )

    period_conversations = list(
        (
            await db.execute(
                select(Conversation)
                .where(*base, Conversation.created_at >= since)
                .order_by(Conversation.created_at.desc())
                .limit(limit * 3)
            )
        ).scalars().all()
    )

    with_first_response = [c for c in period_conversations if c.first_response_at]
    avg_first_response_minutes = None
    if with_first_response:
        deltas = [
            (c.first_response_at - c.created_at).total_seconds() / 60
            for c in with_first_response
            if c.first_response_at and c.created_at
        ]
        avg_first_response_minutes = round(sum(deltas) / len(deltas), 1) if deltas else None

    closed_in_period = [c for c in period_conversations if c.closed_at]
    avg_resolution_minutes = None
    if closed_in_period:
        deltas = [
            (c.closed_at - c.created_at).total_seconds() / 60
            for c in closed_in_period
            if c.closed_at and c.created_at
        ]
        avg_resolution_minutes = round(sum(deltas) / len(deltas), 1) if deltas else None

    now = datetime.now(UTC)
    sla_breach_count = 0
    sla_breaches: list[dict] = []
    for conv in period_conversations:
        breached = False
        response_minutes = None
        if conv.first_response_at and conv.created_at:
            response_minutes = round((conv.first_response_at - conv.created_at).total_seconds() / 60, 1)
            breached = (conv.first_response_at - conv.created_at) > sla_delta
        elif conv.status in {ConversationStatus.OPEN, ConversationStatus.PENDING} and conv.created_at:
            breached = (now - conv.created_at) > sla_delta
            response_minutes = round((now - conv.created_at).total_seconds() / 60, 1)
        if not breached:
            continue
        sla_breach_count += 1
        if limit <= 0 or len(sla_breaches) >= limit:
            continue
        contact = await db.get(Contact, conv.contact_id)
        if contact is None:
            continue
        sla_breaches.append({
            **_contact_row(contact),
            "conversation_id": str(conv.id),
            "status": conv.status.value if hasattr(conv.status, "value") else str(conv.status),
            "waiting_minutes": response_minutes,
            "created_at": conv.created_at.isoformat() if conv.created_at else None,
        })

    return {
        "period_days": days,
        "sla_target_minutes": sla_minutes,
        "summary": {
            "open": open_count,
            "pending": pending_count,
            "closed": closed_count,
            "created_in_period": period_count,
            "sla_breaches": sla_breach_count,
            "avg_first_response_minutes": avg_first_response_minutes,
            "avg_resolution_minutes": avg_resolution_minutes,
        },
        "sla_breaches": sla_breaches,
    }


async def inactivity_report(
    db: AsyncSession, *, account_id: UUID, inactive_days: int = 30, limit: int = 50
) -> dict:
    from sqlalchemy import and_, or_

    threshold = datetime.now(UTC) - timedelta(days=inactive_days)

    last_message_subq = (
        select(
            Message.contact_id.label("contact_id"),
            func.max(Message.created_at).label("last_message_at"),
        )
        .where(Message.account_id == account_id, Message.contact_id.is_not(None))
        .group_by(Message.contact_id)
        .subquery()
    )

    never_messaged_count = int(
        (await db.scalar(
            select(func.count(Contact.id))
            .outerjoin(last_message_subq, Contact.id == last_message_subq.c.contact_id)
            .where(
                Contact.account_id == account_id,
                Contact.deleted_at.is_(None),
                last_message_subq.c.last_message_at.is_(None),
                Contact.created_at <= threshold,
            )
        ))
        or 0
    )
    dormant_count = int(
        (await db.scalar(
            select(func.count(Contact.id))
            .join(last_message_subq, Contact.id == last_message_subq.c.contact_id)
            .where(
                Contact.account_id == account_id,
                Contact.deleted_at.is_(None),
                last_message_subq.c.last_message_at < threshold,
            )
        ))
        or 0
    )

    result = await db.execute(
        select(Contact, last_message_subq.c.last_message_at)
        .outerjoin(last_message_subq, Contact.id == last_message_subq.c.contact_id)
        .where(
            Contact.account_id == account_id,
            Contact.deleted_at.is_(None),
            or_(
                and_(last_message_subq.c.last_message_at.is_(None), Contact.created_at <= threshold),
                last_message_subq.c.last_message_at < threshold,
            ),
        )
        .order_by(last_message_subq.c.last_message_at.asc().nullsfirst(), Contact.created_at.asc())
        .limit(limit * 2)
    )

    never_messaged: list[dict] = []
    dormant: list[dict] = []
    for contact, last_message_at in result.all():
        row = {
            **_contact_row(contact),
            "last_message_at": last_message_at.isoformat() if last_message_at else None,
            "inactive_days": inactive_days,
        }
        if last_message_at is None:
            never_messaged.append(row)
        else:
            dormant.append(row)

    return {
        "inactive_days": inactive_days,
        "summary": {
            "never_messaged": never_messaged_count,
            "dormant": dormant_count,
            "total_inactive": never_messaged_count + dormant_count,
        },
        "never_messaged": never_messaged[:limit],
        "dormant": dormant[:limit],
    }


async def catalog_report(db: AsyncSession, *, account_id: UUID, limit: int = 50) -> dict:
    products = list(
        (
            await db.execute(
                select(CatalogProduct)
                .where(CatalogProduct.account_id == account_id, CatalogProduct.is_active.is_(True))
                .order_by(CatalogProduct.name.asc())
            )
        ).scalars().all()
    )

    def product_row(item: CatalogProduct) -> dict:
        return {
            "id": str(item.id),
            "name": item.name,
            "product_type": item.product_type,
            "price": str(item.price) if item.price is not None else None,
            "price_type": item.price_type,
            "currency": item.currency,
            "has_description": bool(item.description and item.description.strip()),
            "keywords": item.keywords,
        }

    without_price = [
        p for p in products
        if p.price_type == "quote" or p.price is None
    ]
    without_description = [p for p in products if not (p.description and p.description.strip())]
    without_image = [p for p in products if not (p.image_url and p.image_url.strip())]
    services = [p for p in products if p.product_type == "service"]
    items = [p for p in products if p.product_type == "product"]

    return {
        "summary": {
            "total": len(products),
            "products": len(items),
            "services": len(services),
            "without_price": len(without_price),
            "without_description": len(without_description),
            "without_image": len(without_image),
            "without_retailer_id": len([p for p in products if not (p.meta_retailer_id and p.meta_retailer_id.strip())]),
        },
        "without_price": [product_row(p) for p in without_price[:limit]],
        "without_description": [product_row(p) for p in without_description[:limit]],
        "without_image": [product_row(p) for p in without_image[:limit]],
        "top_used": [
            {
                "id": str(item.id),
                "name": item.name,
                "usage_count": item.usage_count,
            }
            for item in sorted(products, key=lambda p: p.usage_count, reverse=True)[:limit]
            if item.usage_count > 0
        ],
        "all_items": [product_row(p) for p in products[:limit]],
    }


async def knowledge_report(db: AsyncSession, *, account_id: UUID, limit: int = 50) -> dict:
    from app.models.knowledge_article import KnowledgeArticle

    articles = list(
        (
            await db.execute(
                select(KnowledgeArticle)
                .where(KnowledgeArticle.account_id == account_id, KnowledgeArticle.is_active.is_(True))
                .order_by(KnowledgeArticle.title.asc())
            )
        ).scalars().all()
    )

    def article_row(item: KnowledgeArticle) -> dict:
        return {
            "id": str(item.id),
            "title": item.title,
            "category": item.category,
            "keywords": item.keywords,
            "usage_count": item.usage_count,
            "language": item.language,
        }

    unused = [a for a in articles if (a.usage_count or 0) == 0]
    top_used = sorted(articles, key=lambda a: a.usage_count or 0, reverse=True)

    return {
        "summary": {
            "total": len(articles),
            "unused": len(unused),
            "total_usage": sum(a.usage_count or 0 for a in articles),
        },
        "top_used": [article_row(a) for a in top_used if a.usage_count > 0][:limit],
        "unused": [article_row(a) for a in unused[:limit]],
        "by_category": [
            {"category": cat, "count": len([a for a in articles if a.category == cat])}
            for cat in sorted({a.category for a in articles})
        ],
    }


async def quick_replies_report(db: AsyncSession, *, account_id: UUID, limit: int = 50) -> dict:
    from app.services.quick_replies import quick_replies_report as _report

    return await _report(db, account_id=account_id, limit=limit)


async def export_report_csv(db: AsyncSession, *, account_id: UUID, report_type: str, days: int = 30) -> str:
    import csv
    import io

    buffer = io.StringIO()
    writer = csv.writer(buffer)

    if report_type == "customers":
        data = await customer_report(db, account_id=account_id, days=days, limit=500)
        writer.writerow(["name", "phone", "email", "country_code", "created_at"])
        for item in data["recent_contacts"]:
            writer.writerow([
                item["display_name"] or "",
                item["phone"],
                item["email"] or "",
                item["country_code"] or "",
                item["created_at"] or "",
            ])
    elif report_type == "names":
        data = await names_report(db, account_id=account_id, limit=500)
        writer.writerow(["report_section", "name", "phone", "email", "created_at"])
        for item in data["missing_names"]:
            writer.writerow(["missing_name", "", item["phone"], item["email"] or "", item["created_at"] or ""])
        for item in data["named_contacts"]:
            writer.writerow(["named", item["display_name"] or "", item["phone"], item["email"] or "", item["created_at"] or ""])
        for item in data["duplicate_names"]:
            writer.writerow(["duplicate", item["name"], "", "", item["count"]])
    elif report_type == "engagement":
        data = await engagement_report(db, account_id=account_id, days=days, limit=500)
        writer.writerow(["status", "name", "phone", "email", "created_at"])
        sections = [
            ("two_way_engaged", data["two_way_engaged"]),
            ("waiting_team_reply", data["waiting_team_reply"]),
            ("waiting_customer_reply", data["waiting_customer_reply"]),
            ("no_interaction", data["no_interaction"]),
        ]
        for label, rows in sections:
            for item in rows:
                writer.writerow([
                    label,
                    item["display_name"] or "",
                    item["phone"],
                    item["email"] or "",
                    item["created_at"] or "",
                ])
    elif report_type == "overview":
        data = await reports_overview(db, account_id=account_id, days=days)
        writer.writerow(["metric", "value"])
        for key, value in data.items():
            writer.writerow([key, value])
    elif report_type == "campaigns":
        data = await campaigns_report(db, account_id=account_id, days=days, limit=500)
        writer.writerow(["name", "status", "total", "sent", "delivered", "read", "failed", "delivery_rate", "read_rate", "created_at"])
        for item in data["campaigns"]:
            writer.writerow([
                item["name"], item["status"], item["total"], item["sent"], item["delivered"],
                item["read"], item["failed"], item["delivery_rate"], item["read_rate"], item["created_at"],
            ])
    elif report_type == "conversations":
        data = await conversations_report(db, account_id=account_id, days=days, limit=500)
        writer.writerow(["name", "phone", "status", "waiting_minutes", "created_at"])
        for item in data["sla_breaches"]:
            writer.writerow([
                item.get("display_name") or "", item.get("phone") or "", item.get("status") or "",
                item.get("waiting_minutes") or "", item.get("created_at") or "",
            ])
    elif report_type == "inactivity":
        data = await inactivity_report(db, account_id=account_id, inactive_days=days, limit=500)
        writer.writerow(["section", "name", "phone", "email", "last_message_at"])
        for item in data["never_messaged"]:
            writer.writerow(["never_messaged", item.get("display_name") or "", item.get("phone") or "", item.get("email") or "", ""])
        for item in data["dormant"]:
            writer.writerow(["dormant", item.get("display_name") or "", item.get("phone") or "", item.get("email") or "", item.get("last_message_at") or ""])
    elif report_type == "catalog":
        data = await catalog_report(db, account_id=account_id, limit=500)
        writer.writerow(["name", "type", "price", "price_type", "currency", "keywords"])
        for item in data["all_items"]:
            writer.writerow([item["name"], item["product_type"], item["price"] or "", item["price_type"], item["currency"], item["keywords"] or ""])
    elif report_type == "knowledge":
        data = await knowledge_report(db, account_id=account_id, limit=500)
        writer.writerow(["title", "category", "keywords", "usage_count", "language"])
        for item in data["top_used"] + data["unused"]:
            writer.writerow([item["title"], item["category"], item["keywords"] or "", item["usage_count"], item["language"]])
    elif report_type == "quick_replies":
        data = await quick_replies_report(db, account_id=account_id, limit=500)
        writer.writerow(["shortcut", "title", "category", "usage_count", "tags"])
        for item in data["top_used"] + data["unused"]:
            writer.writerow([item["shortcut"], item["title"], item.get("category") or "", item["usage_count"], item.get("tags") or ""])
    elif report_type in {"compliance", "team", "automations", "whatsapp", "executive", "audit", "roi"}:
        from app.services.reports_extended import (
            audit_report,
            automations_report,
            campaign_roi_report,
            compliance_report,
            executive_summary_report,
            team_report,
            whatsapp_ops_report,
        )

        if report_type == "compliance":
            data = await compliance_report(db, account_id=account_id, limit=500)
            writer.writerow(["section", "phone", "name", "email", "count"])
            for item in data["opt_out_contacts"]:
                writer.writerow(["opt_out", item["phone"], item.get("display_name") or "", item.get("email") or "", ""])
            for item in data["duplicate_phones"]:
                writer.writerow(["duplicate", item["phone"], "", "", item["count"]])
        elif report_type == "team":
            data = await team_report(db, account_id=account_id, days=days)
            writer.writerow(["agent", "role", "open", "closed", "first_response_min", "sla_pct", "csat", "deals_won"])
            for item in data["agents"]:
                writer.writerow([
                    item["user_name"], item["role"], item["open_conversations"], item["closed_conversations"],
                    item.get("first_response_avg_minutes") or "", item.get("sla_compliance_pct") or "",
                    item.get("csat_average") or "", item.get("deals_won") or 0,
                ])
        elif report_type == "automations":
            data = await automations_report(db, account_id=account_id, days=days, limit=500)
            writer.writerow(["name", "status", "trigger", "runs", "succeeded", "failed", "success_rate"])
            for item in data["automations"]:
                writer.writerow([
                    item["name"], item["status"], item["trigger_type"],
                    item["runs"], item["succeeded"], item["failed"], item.get("success_rate") or "",
                ])
        elif report_type == "whatsapp":
            data = await whatsapp_ops_report(db, account_id=account_id)
            writer.writerow(["phone", "name", "status", "quality", "tier"])
            for item in data["accounts"]:
                writer.writerow([
                    item["display_phone_number"], item.get("verified_name") or "",
                    item["status"], item.get("quality_rating") or "", item.get("messaging_limit_tier") or "",
                ])
        elif report_type == "executive":
            data = await executive_summary_report(db, account_id=account_id, days=days)
            writer.writerow(["section", "metric", "value"])
            for key, value in data["overview"].items():
                if key != "changes_pct":
                    writer.writerow(["overview", key, value])
            for key, value in data["crm"].items():
                writer.writerow(["crm", key, value])
        elif report_type == "audit":
            data = await audit_report(db, account_id=account_id, days=days, limit=500)
            writer.writerow(["action", "resource_type", "resource_id", "actor", "created_at"])
            for item in data["events"]:
                writer.writerow([
                    item["action"], item["resource_type"], item.get("resource_id") or "",
                    item.get("actor_name") or "", item.get("created_at") or "",
                ])
        elif report_type == "roi":
            data = await campaign_roi_report(db, account_id=account_id, days=days)
            writer.writerow(["metric", "value"])
            for key, value in data["summary"].items():
                writer.writerow([key, value])
    else:
        raise ValueError("UNKNOWN_REPORT_TYPE")

    return buffer.getvalue()


def _workbook_to_bytes(workbook) -> bytes:
    import io

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _write_summary_sheet(sheet, rows: list[tuple[str, object]]) -> None:
    sheet.append(["المؤشر", "القيمة"])
    for label, value in rows:
        sheet.append([label, value])


def _write_contacts_sheet(
    sheet,
    rows: list[dict],
    *,
    columns: list[tuple[str, str]] | None = None,
) -> None:
    cols = columns or [
        ("display_name", "الاسم"),
        ("phone", "الرقم"),
        ("email", "البريد"),
        ("country_code", "الدولة"),
        ("created_at", "تاريخ الإضافة"),
    ]
    sheet.append([label for _, label in cols])
    for item in rows:
        sheet.append([item.get(key) or "" for key, _ in cols])


async def export_report_xlsx(db: AsyncSession, *, account_id: UUID, report_type: str, days: int = 30) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    workbook.remove(workbook.active)

    if report_type == "customers":
        data = await customer_report(db, account_id=account_id, days=days, limit=500)
        summary = workbook.create_sheet("ملخص")
        _write_summary_sheet(summary, [
            ("الفترة (أيام)", data["period_days"]),
            ("إجمالي العملاء", data["summary"]["total_contacts"]),
            ("عملاء جدد", data["summary"]["new_contacts"]),
            ("لديهم بريد", data["summary"]["with_email"]),
            ("لديهم محادثات", data["summary"]["with_conversations"]),
            ("بدون محادثات", data["summary"]["without_conversations"]),
        ])
        by_country = workbook.create_sheet("حسب الدولة")
        by_country.append(["الدولة", "العدد"])
        for row in data["by_country"]:
            by_country.append([row["country_code"], row["count"]])
        contacts = workbook.create_sheet("العملاء")
        _write_contacts_sheet(contacts, data["recent_contacts"])
    elif report_type == "names":
        data = await names_report(db, account_id=account_id, limit=500)
        summary = workbook.create_sheet("ملخص")
        _write_summary_sheet(summary, [
            ("إجمالي العملاء", data["summary"]["total_contacts"]),
            ("لديهم اسم", data["summary"]["with_name"]),
            ("بدون اسم", data["summary"]["without_name"]),
            ("مجموعات أسماء مكررة", data["summary"]["duplicate_name_groups"]),
        ])
        duplicates = workbook.create_sheet("أسماء مكررة")
        duplicates.append(["الاسم", "التكرار"])
        for row in data["duplicate_names"]:
            duplicates.append([row["name"], row["count"]])
        missing = workbook.create_sheet("بدون اسم")
        _write_contacts_sheet(
            missing,
            data["missing_names"],
            columns=[
                ("phone", "الرقم"),
                ("email", "البريد"),
                ("created_at", "تاريخ الإضافة"),
            ],
        )
        named = workbook.create_sheet("قائمة الأسماء")
        _write_contacts_sheet(named, data["named_contacts"])
    elif report_type == "engagement":
        data = await engagement_report(db, account_id=account_id, days=days, limit=500)
        summary = workbook.create_sheet("ملخص")
        _write_summary_sheet(summary, [
            ("الفترة (أيام)", data["period_days"]),
            ("تجاوب ثنائي", data["summary"]["two_way_engaged"]),
            ("بانتظار رد الفريق", data["summary"]["waiting_team_reply"]),
            ("بانتظار رد العميل", data["summary"]["waiting_customer_reply"]),
            ("بدون تفاعل", data["summary"]["no_interaction"]),
            ("محادثات مفتوحة بانتظار الفريق", data["summary"]["open_conversations_waiting_team"]),
        ])
        sections = [
            ("تجاوب ثنائي", data["two_way_engaged"]),
            ("بانتظار رد الفريق", data["waiting_team_reply"]),
            ("بانتظار رد العميل", data["waiting_customer_reply"]),
            ("بدون تفاعل", data["no_interaction"]),
        ]
        for sheet_name, rows in sections:
            sheet = workbook.create_sheet(sheet_name[:31])
            _write_contacts_sheet(
                sheet,
                rows,
                columns=[
                    ("display_name", "الاسم"),
                    ("phone", "الرقم"),
                    ("email", "البريد"),
                    ("created_at", "تاريخ الإضافة"),
                ],
            )
    elif report_type == "overview":
        data = await reports_overview(db, account_id=account_id, days=days)
        summary = workbook.create_sheet("نظرة عامة")
        _write_summary_sheet(summary, [
            ("الفترة (أيام)", data["period_days"]),
            ("إجمالي العملاء", data["total_contacts"]),
            ("عملاء جدد", data["new_contacts"]),
            ("لديهم اسم", data["contacts_with_name"]),
            ("بدون اسم", data["contacts_without_name"]),
            ("تجاوب ثنائي", data["two_way_engaged"]),
            ("بانتظار رد الفريق", data["waiting_team_reply"]),
            ("بانتظار رد العميل", data["waiting_customer_reply"]),
            ("بدون تفاعل", data["no_interaction"]),
            ("حملات في الفترة", data["campaigns_in_period"]),
            ("محادثات مفتوحة", data["open_conversations"]),
            ("تجاوز SLA", data["sla_breaches"]),
            ("عملاء غير نشطين", data["inactive_contacts"]),
        ])
    elif report_type == "campaigns":
        data = await campaigns_report(db, account_id=account_id, days=days, limit=500)
        summary = workbook.create_sheet("ملخص")
        s = data["summary"]
        _write_summary_sheet(summary, [
            ("الفترة (أيام)", data["period_days"]),
            ("عدد الحملات", s["campaigns"]),
            ("إجمالي المستلمين", s["recipients"]),
            ("مرسل", s["sent"]),
            ("مُسلَّم", s["delivered"]),
            ("مقروء", s["read"]),
            ("فشل", s["failed"]),
            ("نسبة التسليم %", s["delivery_rate"]),
            ("نسبة القراءة %", s["read_rate"]),
        ])
        sheet = workbook.create_sheet("الحملات")
        sheet.append(["الاسم", "الحالة", "الإجمالي", "مرسل", "مُسلَّم", "مقروء", "فشل", "تسليم%", "قراءة%"])
        for item in data["campaigns"]:
            sheet.append([
                item["name"], item["status"], item["total"], item["sent"], item["delivered"],
                item["read"], item["failed"], item["delivery_rate"], item["read_rate"],
            ])
    elif report_type == "conversations":
        data = await conversations_report(db, account_id=account_id, days=days, limit=500)
        summary = workbook.create_sheet("ملخص")
        s = data["summary"]
        _write_summary_sheet(summary, [
            ("الفترة (أيام)", data["period_days"]),
            ("SLA (دقائق)", data["sla_target_minutes"]),
            ("مفتوحة", s["open"]),
            ("معلّقة", s["pending"]),
            ("مغلقة", s["closed"]),
            ("جديدة في الفترة", s["created_in_period"]),
            ("تجاوز SLA", s["sla_breaches"]),
            ("متوسط أول رد (د)", s["avg_first_response_minutes"]),
            ("متوسط الإغلاق (د)", s["avg_resolution_minutes"]),
        ])
        sheet = workbook.create_sheet("تجاوز SLA")
        sheet.append(["الاسم", "الرقم", "الحالة", "انتظار (د)", "تاريخ المحادثة"])
        for item in data["sla_breaches"]:
            sheet.append([
                item.get("display_name") or "", item.get("phone") or "", item.get("status") or "",
                item.get("waiting_minutes") or "", item.get("created_at") or "",
            ])
    elif report_type == "inactivity":
        data = await inactivity_report(db, account_id=account_id, inactive_days=days, limit=500)
        summary = workbook.create_sheet("ملخص")
        s = data["summary"]
        _write_summary_sheet(summary, [
            ("أيام عدم النشاط", data["inactive_days"]),
            ("لم يراسلوا أبداً", s["never_messaged"]),
            ("خاملين", s["dormant"]),
            ("إجمالي غير نشط", s["total_inactive"]),
        ])
        never = workbook.create_sheet("لم يراسلوا")
        _write_contacts_sheet(never, data["never_messaged"], columns=[
            ("display_name", "الاسم"), ("phone", "الرقم"), ("email", "البريد"), ("created_at", "تاريخ الإضافة"),
        ])
        dormant = workbook.create_sheet("خاملين")
        dormant.append(["الاسم", "الرقم", "البريد", "آخر رسالة"])
        for item in data["dormant"]:
            dormant.append([item.get("display_name") or "", item.get("phone") or "", item.get("email") or "", item.get("last_message_at") or ""])
    elif report_type == "catalog":
        data = await catalog_report(db, account_id=account_id, limit=500)
        summary = workbook.create_sheet("ملخص")
        s = data["summary"]
        _write_summary_sheet(summary, [
            ("إجمالي", s["total"]),
            ("منتجات", s["products"]),
            ("خدمات", s["services"]),
            ("بدون سعر", s["without_price"]),
            ("بدون وصف", s["without_description"]),
        ])
        sheet = workbook.create_sheet("الكتalog")
        sheet.append(["الاسم", "النوع", "السعر", "نوع السعر", "العملة"])
        for item in data["all_items"]:
            sheet.append([item["name"], item["product_type"], item["price"] or "", item["price_type"], item["currency"]])
    elif report_type == "knowledge":
        data = await knowledge_report(db, account_id=account_id, limit=500)
        summary = workbook.create_sheet("ملخص")
        s = data["summary"]
        _write_summary_sheet(summary, [
            ("إجمالي المقالات", s["total"]),
            ("بدون استخدام", s["unused"]),
            ("إجمالي الاستخدام", s["total_usage"]),
        ])
        sheet = workbook.create_sheet("الأكثر استخداماً")
        sheet.append(["العنوان", "الفئة", "الاستخدام"])
        for item in data["top_used"]:
            sheet.append([item["title"], item["category"], item["usage_count"]])
    elif report_type == "quick_replies":
        data = await quick_replies_report(db, account_id=account_id, limit=500)
        summary = workbook.create_sheet("ملخص")
        s = data["summary"]
        _write_summary_sheet(summary, [
            ("إجمالي الردود", s["total"]),
            ("بدون استخدام", s["unused"]),
            ("إجمالي الاستخدام", s["total_usage"]),
        ])
        sheet = workbook.create_sheet("الأكثر استخداماً")
        sheet.append(["الاختصار", "العنوان", "الفئة", "الاستخدام"])
        for item in data["top_used"]:
            sheet.append([item["shortcut"], item["title"], item.get("category") or "", item["usage_count"]])
    elif report_type in {"compliance", "team", "automations", "whatsapp", "executive", "audit", "roi"}:
        from app.services.reports_extended import (
            audit_report,
            automations_report,
            campaign_roi_report,
            compliance_report,
            executive_summary_report,
            team_report,
            whatsapp_ops_report,
        )

        if report_type == "compliance":
            data = await compliance_report(db, account_id=account_id, limit=500)
            summary = workbook.create_sheet("ملخص")
            _write_summary_sheet(summary, [(k, v) for k, v in data["summary"].items()])
            sheet = workbook.create_sheet("رفض التسويق")
            sheet.append(["الاسم", "الرقم", "البريد"])
            for item in data["opt_out_contacts"]:
                sheet.append([item.get("display_name") or "", item["phone"], item.get("email") or ""])
        elif report_type == "team":
            data = await team_report(db, account_id=account_id, days=days)
            sheet = workbook.create_sheet("الفريق")
            sheet.append(["الموظف", "الدور", "مفتوحة", "مغلقة", "أول رد", "SLA%", "CSAT", "فوز"])
            for item in data["agents"]:
                sheet.append([
                    item["user_name"], item["role"], item["open_conversations"], item["closed_conversations"],
                    item.get("first_response_avg_minutes") or "", item.get("sla_compliance_pct") or "",
                    item.get("csat_average") or "", item.get("deals_won") or 0,
                ])
        elif report_type == "automations":
            data = await automations_report(db, account_id=account_id, days=days, limit=500)
            sheet = workbook.create_sheet("أتمتة")
            sheet.append(["الاسم", "الحالة", "المشغّل", "تشغيلات", "نجاح", "فشل", "نسبة"])
            for item in data["automations"]:
                sheet.append([
                    item["name"], item["status"], item["trigger_type"],
                    item["runs"], item["succeeded"], item["failed"], item.get("success_rate") or "",
                ])
        elif report_type == "whatsapp":
            data = await whatsapp_ops_report(db, account_id=account_id)
            sheet = workbook.create_sheet("WhatsApp")
            sheet.append(["الرقم", "الاسم", "الحالة", "الجودة", "الحد"])
            for item in data["accounts"]:
                sheet.append([
                    item["display_phone_number"], item.get("verified_name") or "",
                    item["status"], item.get("quality_rating") or "", item.get("messaging_limit_tier") or "",
                ])
        elif report_type == "executive":
            data = await executive_summary_report(db, account_id=account_id, days=days)
            sheet = workbook.create_sheet("Executive")
            _write_summary_sheet(sheet, [
                ("الفترة", data["period_days"]),
                ("Pipeline", data["crm"]["pipeline_value"]),
                ("فوز الشهر", data["crm"]["won_value_month"]),
                ("CSAT", data["csat"].get("average_score")),
                ("SLA avg", data["sla"].get("first_response_avg_minutes")),
            ])
        elif report_type == "audit":
            data = await audit_report(db, account_id=account_id, days=days, limit=500)
            sheet = workbook.create_sheet("Audit")
            sheet.append(["الإجراء", "النوع", "المعرف", "المستخدم", "التاريخ"])
            for item in data["events"]:
                sheet.append([
                    item["action"], item["resource_type"], item.get("resource_id") or "",
                    item.get("actor_name") or "", item.get("created_at") or "",
                ])
        elif report_type == "roi":
            data = await campaign_roi_report(db, account_id=account_id, days=days)
            summary = workbook.create_sheet("ROI")
            _write_summary_sheet(summary, [(k, v) for k, v in data["summary"].items()])
    else:
        raise ValueError("UNKNOWN_REPORT_TYPE")

    return _workbook_to_bytes(workbook)
